import FDPClient
import Dataset
import Organisation
import Biobank
import Patientregistry
import Config
import chevron
import openpyxl
import csv
import Utils
import Distribution
from rdflib import Graph



class Populator:
    """
    Class contents methods to extract content from the input CSV files and methods to populate FDP with content.
    """
    FDP_CLIENT = FDPClient.FDPClient(Config.FDP_URL, Config.FDP_USERNAME, Config.FDP_PASSWORD,
                                     Config.FDP_PERSISTENT_URL)
    UTILS = Utils.Utils()

    def __init__(self):
        """
        This __init__ method exacts datasets and distribution objects from the input CSV files. These objects are used to
        create metadata entries in the FAIR Data Point.
        """
        # # GET datasets
        # datasets = self.__get_datasets__()
        # # GET distributions
        # distributions = self.__get_distributions__()
        # # Populate FDP with datasets
        # for dataset_name, dataset in datasets.items():
        #     dataset_url = self.create_dataset(dataset)
        #     # Populate FDP with distribution(s) as child to dataset
        #     for distribution_name, distribution in distributions.items():
        #         if distribution.DATASET_NAME == dataset_name:
        #             distribution.PARENT_URL = dataset_url

        #             # This logic is required since both download and access URLs are captured in same row
        #             download_url = distribution.DOWNLOAD_URL
        #             distribution_name = distribution.TITLE
        #             if distribution.ACCESS_URL:
        #                 distribution.TITLE = "Access distribution of : " + distribution_name
        #                 distribution.DOWNLOAD_URL = None
        #                 self.create_distribution(distribution)

        #             if download_url:
        #                 distribution.TITLE = "Downloadable distribution of : " + distribution_name
        #                 distribution.ACCESS_URL = None
        #                 distribution.DOWNLOAD_URL = download_url
        #                 self.create_distribution(distribution)

        organisations = self.__get_organisations__()
        biobanks = self.__get_biobanks__()
        patientregistries = self.__get_patientregistries__()
        print(patientregistries)

        # biobanks = [Biobank.Biobank(Config.CATALOG_URL, None, "Biobank test",
        #     "Test of biobank pushed to FDP using FDP populator.", "National",
        #     ["https://example.org/ont/example", "https://example.org/ont/example2"], "Biobank organisation",
        #     ["https://example.org/biobank", "https://example.org/extra_page"])]

        # organisations = [Organisation.Organisation(Config.CATALOG_URL, "Biobank organisation",
        #     "This is an organisation pushed to FDP using FDP populator", "The Netherlands", 
        #     "Leiden", ["https://example.org/biobankorganisation"])]

        for organisation_name, organisation in organisations.items():
            organisation_url = self.create_organisation(organisation)
            for biobank_name, biobank in biobanks.items():
                if biobank.PUBLISHER_NAME == organisation.TITLE:
                    biobank.PUBLISHER_URL = organisation_url
                    self.create_biobank(biobank)
            for patientregistry_name, patientregistry in patientregistries.items():
                if patientregistry.PUBLISHER_NAME == organisation.TITLE:
                    patientregistry.PUBLISHER_URL = organisation_url
                    self.create_patientregistry(patientregistry)
        
        # for biobank in biobanks:
            
        # print(biobank)
        # print(organisation)
        # self.create_biobank(biobank)
        # #self.create_organisation(organisation)

    def create_dataset(self, dataset):
        """
        Method to create dataset in FDP

        :param dataset: Provide dataset object
        :return: FDP's dataset URL
        """
        parent_url = dataset.PARENT_URL

        if not self.FDP_CLIENT.does_metadata_exists(parent_url):
            raise SystemExit("The catalog <"+parent_url+"> doesn't exist. Provide valid catalog URL")

        print("The catalog <"+parent_url+"> exist")

        graph = Graph()

        # create resource triples
        self.UTILS.add_resource_triples(dataset, graph)
        # Create language triples
        self.UTILS.add_language_triples(dataset, graph)
        # Create license triples
        self.UTILS.add_licence_triples(dataset, graph)

        # Create landing page triples
        if dataset.LANDING_PAGE:
            with open('../templates/landingpage.mustache', 'r') as f:
                body = chevron.render(f, {'page_url': dataset.LANDING_PAGE})
                graph.parse(data=body, format="turtle")

        # Create contact point triples
        if dataset.CONTACT_POINT:
            with open('../templates/contact.mustache', 'r') as f:
                body = chevron.render(f, {'contact_url': dataset.CONTACT_POINT})
                graph.parse(data=body, format="turtle")

        # Create keywords list
        keyword_str = ""
        for keyword in dataset.KEYWORDS:
            keyword_str = keyword_str + ' "' + keyword + '",'
        keyword_str = keyword_str[:-1]

        # Create themes list
        theme_str = ""
        for theme in dataset.THEMES:
            theme_str = theme_str + " <" + theme + ">,"
        theme_str = theme_str[:-1]

        # create dataset triples
        with open('../templates/dataset.mustache', 'r') as f:
            body = chevron.render(f, {'keyword': keyword_str, 'theme': theme_str})
            graph.parse(data=body, format="turtle")

        post_body = graph.serialize(format='turtle')
        dataset_url = self.FDP_CLIENT.fdp_create_metadata(post_body, "dataset")
        print("New dataset created : " + dataset_url)
        return dataset_url

    def create_distribution(self, distribution):
        """
        Method to create distribution in FDP

        :param distribution: Provide distribution object
        :return: FDP's distribution URL
        """
        parent_url = distribution.PARENT_URL

        if not self.FDP_CLIENT.does_metadata_exists(parent_url):
            raise SystemExit("The dataset <"+parent_url+"> doesn't exist. Provide valid dataset URL")

        graph = Graph()

        # create resource triples
        self.UTILS.add_resource_triples(distribution, graph)
        # Create language triples
        self.UTILS.add_language_triples(distribution, graph)
        # Create license triples
        self.UTILS.add_licence_triples(distribution, graph)

        # Create byte size triples
        if distribution.BYTE_SIZE:
            with open('../templates/bytesize.mustache', 'r') as f:
                body = chevron.render(f, {'byte_size': distribution.BYTE_SIZE})
                graph.parse(data=body, format="turtle")

        # Create format triples
        if distribution.FORMAT:
            with open('../templates/format.mustache', 'r') as f:
                body = chevron.render(f, {'format': distribution.FORMAT})
                graph.parse(data=body, format="turtle")

        distribution_url = None
        distribution_type = None

        if distribution.ACCESS_URL:
            distribution_type = "dcat:accessURL"
            distribution_url = distribution.ACCESS_URL
        elif distribution.DOWNLOAD_URL:
            distribution_type = "dcat:downloadURL"
            distribution_url = distribution.DOWNLOAD_URL

        # create distribution triples
        with open('../templates/distribution.mustache', 'r') as f:
            body = chevron.render(f, {'distribution_type': distribution_type, 'distribution_url': distribution_url,
                                      'media_type': distribution.MEDIA_TYPE})
            graph.parse(data=body, format="turtle")

        post_body = graph.serialize(format='turtle')
        metadata_url = self.FDP_CLIENT.fdp_create_metadata(post_body, "distribution")
        print("New distribution created : " + metadata_url)
        return metadata_url

    def create_organisation(self, organisation):
        """
        Method to create organisation in FDP

        :param biobank: Provide organisation object
        :return: FDP's organisation URL
        """
        parent_url = organisation.PARENT_URL

        if not self.FDP_CLIENT.does_metadata_exists(parent_url):
            raise SystemExit("The catalog <"+parent_url+"> doesn't exist. Provide valid catalog URL")

        print("The catalog <"+parent_url+"> exist")

        # Create pages list
        page_str = ""
        for page in organisation.LANDING_PAGES:
            page_str = page_str + " <" + page + ">,"
        page_str = page_str[:-1]

        # Render RDF
        graph = Graph()

        with open('../templates/organisation.mustache', 'r') as f:
            body = chevron.render(f, {'parent_url': organisation.PARENT_URL,
                                      'title': organisation.TITLE,
                                      'description': organisation.DESCRIPTION,
                                      'location_title': organisation.LOCATION_TITLE,
                                      'location_description': organisation.LOCATION_DESCRIPTION,
                                      'pages': page_str})
            graph.parse(data=body, format="turtle")

        # Serialize RDF and send to FDP
        post_body = graph.serialize(format='turtle')
        print(post_body)
        organisation_url = self.FDP_CLIENT.fdp_create_metadata(post_body, "organisation")
        print("New organisation created : " + organisation_url)
        return organisation_url

    def create_biobank(self, biobank):
        """
        Method to create biobank in FDP

        :param biobank: Provide biobank object
        :return: FDP's biobank URL
        """
        parent_url = biobank.PARENT_URL

        if not self.FDP_CLIENT.does_metadata_exists(parent_url):
            raise SystemExit("The catalog <"+parent_url+"> doesn't exist. Provide valid catalog URL")

        print("The catalog <"+parent_url+"> exist")


        # Create themes list
        theme_str = ""
        for theme in biobank.THEMES:
            theme_str = theme_str + " <" + theme + ">,"
        theme_str = theme_str[:-1]

        # Create pages list
        page_str = ""
        for page in biobank.LANDING_PAGES:
            page_str = page_str + " <" + page + ">,"
        page_str = page_str[:-1]

        # Render RDF
        graph = Graph()

        with open('../templates/biobank.mustache', 'r') as f:
            body = chevron.render(f, {'parent_url': biobank.PARENT_URL,
                                      'title': biobank.TITLE,
                                      'description': biobank.DESCRIPTION,
                                      'populationcoverage': biobank.POPULATIONCOVERAGE,
                                      'themes': theme_str,
                                      'publisher': biobank.PUBLISHER_URL,
                                      'pages': page_str})
            graph.parse(data=body, format="turtle")

        # Serialize RDF and send to FDP
        post_body = graph.serialize(format='turtle')
        print(post_body)
        biobank_url = self.FDP_CLIENT.fdp_create_metadata(post_body, "biobank")
        print("New biobank created : " + biobank_url)
        return biobank_url

    def create_patientregistry(self, patientregistry):
        """
        Method to create patient registry in FDP

        :param patient registry: Provide patient registry object
        :return: FDP's patient registry URL
        """
        parent_url = patientregistry.PARENT_URL

        if not self.FDP_CLIENT.does_metadata_exists(parent_url):
            raise SystemExit("The catalog <"+parent_url+"> doesn't exist. Provide valid catalog URL")

        print("The catalog <"+parent_url+"> exist")


        # Create themes list
        theme_str = ""
        for theme in patientregistry.THEMES:
            theme_str = theme_str + " <" + theme + ">,"
        theme_str = theme_str[:-1]

        # Create pages list
        page_str = ""
        for page in patientregistry.LANDING_PAGES:
            page_str = page_str + " <" + page + ">,"
        page_str = page_str[:-1]

        # Render RDF
        graph = Graph()

        with open('../templates/patientregistry.mustache', 'r') as f:
            body = chevron.render(f, {'parent_url': patientregistry.PARENT_URL,
                                      'title': patientregistry.TITLE,
                                      'description': patientregistry.DESCRIPTION,
                                      'populationcoverage': patientregistry.POPULATIONCOVERAGE,
                                      'themes': theme_str,
                                      'publisher': patientregistry.PUBLISHER_URL,
                                      'pages': page_str})
            graph.parse(data=body, format="turtle")

        # Serialize RDF and send to FDP
        post_body = graph.serialize(format='turtle')
        print(post_body)
        patientregistry_url = self.FDP_CLIENT.fdp_create_metadata(post_body, "patientregistry")
        print("New patient registry created : " + patientregistry_url)
        return patientregistry_url
        
    """
    """
    def __get_datasets__(self):
        """
        This method creates datasets objects by extracting content from the dataset input CSV file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/LUMC-BioSemantics/EJP-RD-WP19-FDP-template>

        :return: Dict of datasets
        """

        reader = csv.reader(open(Config.DATASET_INPUT_FILE, 'r'))
        catalog_url = Config.CATALOG_URL
        datasets = {}
        for row in reader:
            if reader.line_num > 1:
                print(row)
                title = row[0]
                publisher_url = row[1]
                description = row[2]
                language = row[3]
                license = row[4]
                contact_point = row[5]
                landing_page = row[6]
                keywords_str = row[7]
                themes_str = row[8]
                language_url = None
                license_url = None
                landing_page_url = None
                contact_point_url = None

                if not description:
                    description = "Metadata od dataset " + title
                # Create language triples
                if language:
                    language_url = "http://id.loc.gov/vocabulary/iso639-1/" + language.strip()
                # Create license triples
                if license:
                    license_url = license.strip()
                # Create license triples
                if landing_page:
                    landing_page_url = landing_page.strip()
                # Create contact point triples
                if contact_point:
                    contact_point_url = contact_point.strip()
                # Create keywords list
                keywords = []
                for keyword in keywords_str.split(","):
                    keyword = keyword.strip()
                    keywords.append(keyword)
                # Create themes list
                themes = []
                for theme in themes_str.split(","):
                    theme = theme.strip()
                    themes.append(theme)
                dataset = Dataset.Dataset(catalog_url, title, description, keywords, themes, publisher_url,
                                          language_url, license_url, landing_page_url, contact_point_url)
                datasets[title] = dataset
        return datasets

    def __get_distributions__(self):
        """
        This method creates distribution objects by extracting content from the distribution input CSV file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/LUMC-BioSemantics/EJP-RD-WP19-FDP-template>

        :return: Dict of distribution
        """

        reader = csv.reader(open(Config.DISTRIBUTION_INPUT_FILE, 'r'))
        distributions = {}
        for row in reader:
            if reader.line_num > 1:
                print(row)
                title = row[0]
                dataset_name = row[1]
                publisher_url = row[2]
                description = row[3]
                language = row[4]
                license = row[5]
                access_url = row[6]
                download_url = row[7]
                media_type = row[8]
                compression_format = row[9]
                format = row[10]
                byte_size = row[11]
                language_url = None
                license_url = None

                if not description:
                    description = "Metadata od dataset " + title
                # Create language triples
                if language:
                    language_url = "http://id.loc.gov/vocabulary/iso639-1/" + language.strip()
                # Create license triples
                if license:
                    license_url = license.strip()
                distribution = Distribution.Distribution(None, title, description, publisher_url, language_url,
                                                         license_url, access_url, download_url, media_type,
                                                         compression_format, format, byte_size, dataset_name)
                distributions[title] = distribution
        return distributions
    
    def __get_organisations__(self):
        """
        This method creates organisation objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of organisations
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['Organisation']
        
        # Loop over rows of excel sheet
        first_row = True
        organisations = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value

                pages = []
                for page in row[2].value.split(";"):
                    page = page.strip()
                    pages.append(page)

                location_title = row[3].value
                location_description = row[4].value

                # Create organisation object and add to organisation dictionary
                organisation = Organisation.Organisation(Config.CATALOG_URL, title, description, location_title, location_description, pages)
                organisations[organisation.TITLE] = organisation

        return organisations

    def __get_biobanks__(self):
        """
        This method creates biobank objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of biobanks
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        biobanks = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                themes = []
                for theme in row[3].value.split(";"):
                    theme = theme.strip()
                    themes.append(theme)

                publisher_name = row[4].value

                pages = []
                for page in row[5].value.split(";"):
                    page = page.strip()
                    pages.append(page)

                resource_type = row[6].value

                # Create biobank object and add to biobank dictionary if it is a biobank
                if resource_type == "Biobank":
                    biobank = Biobank.Biobank(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    biobanks[biobank.TITLE] = biobank

        return biobanks

    def __get_patientregistries__(self):
        """
        This method creates patient registry objects by extracting content from the ejp vp input file.
        NOTE: This method assumes that provided input file follows this spec
        <https://github.com/ejp-rd-vp/resource-metadata-schema/blob/master/template/EJPRD%20Resource%20Metadata%20template.xlsx>

        :return: Dict of patientregistries
        """
        # Open organisation excel sheet
        wb = openpyxl.load_workbook(Config.EJP_VP_INPUT_FILE)
        ws = wb['BiobankPatientRegistry']
        
        # Loop over rows of excel sheet
        first_row = True
        patientregistries = {}
        for row in ws:
            # Skip header
            if first_row:
                first_row=False
                continue

            if row[0].value != None:
                # Retrieve field values from excel files
                title = row[0].value
                description = row[1].value
                populationcoverage = row[2].value

                themes = []
                for theme in row[3].value.split(";"):
                    theme = theme.strip()
                    themes.append(theme)

                publisher_name = row[4].value

                pages = []
                for page in row[5].value.split(";"):
                    page = page.strip()
                    pages.append(page)

                resource_type = row[6].value

                # Create patient registry object and add to patientregistry dictionary if it is a patientregistry
                if resource_type == "Patient registry":
                    patientregistry = Patientregistry.Patientregistry(Config.CATALOG_URL, None, title, description, populationcoverage, themes, publisher_name, pages)
                    patientregistries[patientregistry.TITLE] = patientregistry

        return patientregistries